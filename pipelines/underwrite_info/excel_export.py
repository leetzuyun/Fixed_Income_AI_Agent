"""
pipelines/underwrite_info/excel_export.py
把一批 records 輸出成 Excel（明細 + 彙總 + 四張 star schema 資料模型
表）。邏輯搬自你原本的 bond_scraper.py 的 write_excel() /
build_model_tables() / write_model_sheet()，格式完全不變——這份 Excel
是給人工瀏覽、分享用的，真正給 agent 查詢用的是
domains/underwriting_kb/store.py 的 SQLite。
"""
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="1F497D")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
SUMMARY_FILL = PatternFill("solid", fgColor="E2EFDA")
SUMMARY_FONT = Font(bold=True, name="Arial", size=10)
DIM_FILL = PatternFill("solid", fgColor="D9E1F2")
FACT_FILL = PatternFill("solid", fgColor="FCE4D6")
NOTE_FONT = Font(italic=True, color="595959", name="Arial", size=9)
NORMAL_FONT = Font(name="Arial", size=10)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _set_cell(ws, row, col, value, font=None, fill=None, alignment=None, border=True):
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = THIN_BORDER
    return cell


def build_model_tables(records: list) -> dict:
    ann_map, bond_map, uw_map = {}, {}, {}
    facts, ann_rows, bond_rows, uw_rows = [], [], [], []
    ann_seq = bond_seq = uw_seq = fact_seq = 1

    for rec in records:
        ann_no = rec.get("序號", "")
        if ann_no not in ann_map:
            ann_map[ann_no] = ann_seq
            date_str = rec.get("申報日期", "")
            parts = date_str.split("/") if "/" in date_str else []
            year = parts[0] if len(parts) > 0 else ""
            month = parts[1] if len(parts) > 1 else ""
            quarter = f"Q{((int(month) - 1) // 3) + 1}" if month.isdigit() else ""
            ann_rows.append({
                "announcement_id": ann_seq, "announcement_no": ann_no,
                "published_date": date_str, "year": year, "month": month,
                "quarter": quarter, "主辦承銷商": rec.get("主辦承銷商", ""),
                "案件名稱": rec.get("案件名稱", ""), "方式": rec.get("方式", ""),
                "發行性質": rec.get("發行性質", ""), "發行種類": rec.get("發行種類", ""),
            })
            ann_seq += 1

        bond_name = rec.get("債券名稱") or rec.get("案件名稱", "—")
        if bond_name not in bond_map:
            bond_map[bond_name] = bond_seq
            bond_rows.append({
                "bond_id": bond_seq, "bond_name": bond_name,
                "issuer_name": rec.get("發行人", "—"),
                "bond_type": rec.get("發行種類", "—"),
                "maturity_date": rec.get("到期日", "—"),
            })
            bond_seq += 1

        uw_name = rec.get("承銷商名稱", "—")
        if uw_name not in uw_map:
            uw_map[uw_name] = uw_seq
            uw_rows.append({
                "underwriter_id": uw_seq, "underwriter_name": uw_name,
                "underwriter_type": "—",
            })
            uw_seq += 1

        facts.append({
            "underwriting_id": fact_seq, "announcement_id": ann_map[ann_no],
            "bond_id": bond_map[bond_name], "underwriter_id": uw_map[uw_name],
            "amount": rec.get("總承銷/洽商銷售金額", "—"),
            "amount_original": rec.get("總承銷/洽商銷售金額", "—"),
            "currency": rec.get("幣別", "TWD"), "fx_rate": "—",
            "總承銷/洽商銷售數量": rec.get("總承銷/洽商銷售數量", "—"),
        })
        fact_seq += 1

    return {
        "Announcement_Dim": ann_rows, "Bond_Dim": bond_rows,
        "Underwriter_Dim": uw_rows, "Underwriting_Fact": facts,
    }


SHEET_META = {
    "Underwriting_Fact": {
        "title": "【事實表】承銷明細", "fill": FACT_FILL,
        "note": "每筆代表一個承銷商對一支債券的承銷紀錄；FK 對應各維度表的 PK",
        "cols": [
            ("underwriting_id", "承銷明細PK", 16), ("announcement_id", "FK→公告", 12),
            ("bond_id", "FK→債券", 12), ("underwriter_id", "FK→券商", 12),
            ("amount", "承銷金額", 24), ("amount_original", "原始金額", 24),
            ("currency", "幣別", 10), ("fx_rate", "匯率", 10),
            ("總承銷/洽商銷售數量", "承銷數量", 20),
        ],
    },
    "Announcement_Dim": {
        "title": "【維度表】公告", "fill": DIM_FILL,
        "note": "每筆代表一份承銷公告；PK = announcement_id",
        "cols": [
            ("announcement_id", "公告PK", 12), ("announcement_no", "序號", 10),
            ("published_date", "申報日期", 14), ("year", "年", 8), ("month", "月", 8),
            ("quarter", "季", 8), ("主辦承銷商", "主辦承銷商", 28),
            ("案件名稱", "案件名稱", 48), ("方式", "方式", 10),
            ("發行性質", "發行性質", 14), ("發行種類", "發行種類", 28),
        ],
    },
    "Bond_Dim": {
        "title": "【維度表】債券", "fill": DIM_FILL,
        "note": "每筆代表一支債券；PK = bond_id",
        "cols": [
            ("bond_id", "債券PK", 10), ("bond_name", "債券名稱", 48),
            ("issuer_name", "發行人", 32), ("bond_type", "債券種類", 28),
            ("maturity_date", "到期日", 16),
        ],
    },
    "Underwriter_Dim": {
        "title": "【維度表】券商", "fill": DIM_FILL,
        "note": "每筆代表一家承銷券商；PK = underwriter_id",
        "cols": [
            ("underwriter_id", "券商PK", 10), ("underwriter_name", "券商名稱", 32),
            ("underwriter_type", "本土/外資", 16),
        ],
    },
}


def write_model_sheet(wb: Workbook, sheet_key: str, data_rows: list):
    meta = SHEET_META[sheet_key]
    ws = wb.create_sheet(meta["title"])
    ws.freeze_panes = "A3"

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(meta["cols"]))
    note_cell = ws.cell(row=1, column=1, value=f"📌 {meta['note']}")
    note_cell.font = NOTE_FONT
    note_cell.alignment = LEFT
    ws.row_dimensions[1].height = 18

    for ci, (_, label, width) in enumerate(meta["cols"], 1):
        _set_cell(ws, 2, ci, label, HEADER_FONT, meta["fill"], CENTER)
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[2].height = 22

    for ri, row in enumerate(data_rows, 3):
        for ci, (key, _, _) in enumerate(meta["cols"], 1):
            val = row.get(key, "")
            align = CENTER if ci == 1 else LEFT
            _set_cell(ws, ri, ci, val, NORMAL_FONT, None, align)


def write_excel(records: list, output_path: str):
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "明細"
    ws1.freeze_panes = "A2"
    headers1 = ["序號", "申報日期", "主辦承銷商", "案件名稱", "方式", "發行性質", "發行種類",
                "承銷商名稱", "總承銷/洽商銷售金額", "總承銷/洽商銷售數量"]
    widths1 = [8, 12, 28, 50, 8, 12, 30, 28, 24, 20]
    for ci, (h, w) in enumerate(zip(headers1, widths1), 1):
        _set_cell(ws1, 1, ci, h, HEADER_FONT, HEADER_FILL, CENTER)
        ws1.column_dimensions[get_column_letter(ci)].width = w
    ws1.row_dimensions[1].height = 24
    for ri, rec in enumerate(records, 2):
        vals = [rec.get(c, "") for c in headers1]
        for ci, val in enumerate(vals, 1):
            align = CENTER if ci in (1, 2, 5) else LEFT
            _set_cell(ws1, ri, ci, val, NORMAL_FONT, None, align)

    ws2 = wb.create_sheet("彙總")
    ws2.freeze_panes = "A2"
    headers2 = ["承銷商名稱", "案件名稱", "序號", "申報日期", "總承銷/洽商銷售金額",
                "總承銷/洽商銷售數量", "備註"]
    widths2 = [28, 50, 8, 12, 24, 20, 20]
    for ci, (h, w) in enumerate(zip(headers2, widths2), 1):
        _set_cell(ws2, 1, ci, h, HEADER_FONT, HEADER_FILL, CENTER)
        ws2.column_dimensions[get_column_letter(ci)].width = w
    ws2.row_dimensions[1].height = 24

    groups = defaultdict(list)
    for rec in records:
        groups[rec.get("承銷商名稱") or "—"].append(rec)

    cur = 2
    for underwriter, grp in sorted(groups.items()):
        grp_start = cur
        for rec in grp:
            vals = ["", rec.get("案件名稱", ""), rec.get("序號", ""),
                    rec.get("申報日期", ""), rec.get("總承銷/洽商銷售金額", "—"),
                    rec.get("總承銷/洽商銷售數量", "—"), ""]
            for ci, val in enumerate(vals, 1):
                align = CENTER if ci in (1, 3, 4) else LEFT
                _set_cell(ws2, cur, ci, val, NORMAL_FONT, None, align)
            cur += 1
        grp_end = cur - 1

        _set_cell(ws2, cur, 1, underwriter, SUMMARY_FONT, SUMMARY_FILL, LEFT)
        _set_cell(ws2, cur, 2, f"【小計：共{len(grp)}筆】", SUMMARY_FONT, SUMMARY_FILL, LEFT)
        for ci in range(3, 8):
            _set_cell(ws2, cur, ci, "", SUMMARY_FONT, SUMMARY_FILL, CENTER)
        cur += 1

        if grp_end >= grp_start:
            ws2.merge_cells(start_row=grp_start, start_column=1, end_row=grp_end, end_column=1)
            cell = ws2.cell(row=grp_start, column=1)
            cell.value = underwriter
            cell.font = NORMAL_FONT
            cell.alignment = LEFT
            cell.border = THIN_BORDER

    model = build_model_tables(records)
    for key in ["Announcement_Dim", "Bond_Dim", "Underwriter_Dim", "Underwriting_Fact"]:
        write_model_sheet(wb, key, model[key])

    wb.save(output_path)
    return output_path
