"""
pipelines/market_report/report1.py
Builds the final PDF report: title page, the two synthesized sections,
optional charts, then embedded source images as an appendix.

搬家後的改動（只有這兩處，其餘邏輯完全不變）：
  1. `import config` -> `import infra.config as config`
  2. _FONT_PATH / _BOLD_FONT_PATH / DEFAULT_BOND_REVIEW_PATH 原本用
     os.path.dirname(os.path.abspath(__file__)) 算路徑，是假設這支檔案
     放在專案根目錄（fonts/、sample_data/ 都在它旁邊）。現在這支檔案搬
     進 pipelines/market_report/ 底下，同樣算法會去找
     pipelines/market_report/fonts/、pipelines/market_report/sample_data/
     這兩個不存在的路徑，字型會 fallback 成 Helvetica（中文變成方塊字），
     Bond Review 圖表也會抓不到資料。改成從 infra/paths.py 拿
     PROJECT_ROOT，維持指向專案根目錄的 fonts/、sample_data/，行為跟搬
     家前一致，不用真的搬動 fonts/、sample_data/ 這兩個資料夾。

Markdown handling
-----------------
The LLM sections come back as Markdown (bullets, **bold**, ## headings,
pipe tables). ReportLab does NOT understand Markdown, so everything below
_render_markdown() exists to turn that text into real flowables:
  "- xxx"        -> Paragraph(bulletText="•")   real hanging-indent bullet
  "1. xxx"       -> numbered bullet
  "## xxx"       -> heading style
  "| a | b |"    -> reportlab Table
  "**bold**"     -> <b>bold</b>
Anything unrecognized falls through as a normal paragraph, so a plain-prose
section (第一區) renders exactly as before.
"""
import io
import glob
import os
import re
from openpyxl import load_workbook

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.graphics.shapes import Drawing, Rect, String, Line
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Image, PageBreak, Table, TableStyle,
    HRFlowable, KeepTogether
)

import infra.config as config
from infra.paths import PROJECT_ROOT

# Cap on how many embedded source images to show per file, so a report with
# 20 exhibit charts doesn't blow the digest up to 20 pages.
MAX_SOURCE_IMAGES_PER_FILE = 4

# --- Minimal report palette ------------------------------------------------
# A restrained navy / blue-grey system keeps the report professional while
# still making dense risk information easy to scan.
NAVY = colors.HexColor("#17324D")
SLATE = colors.HexColor("#52697F")
ACCENT = colors.HexColor("#3E7CB1")
PALE_BLUE = colors.HexColor("#EEF4F8")
PALE_GREY = colors.HexColor("#F7F9FB")
BORDER = colors.HexColor("#D8E1E8")
INK = colors.HexColor("#243746")
MUTED = colors.HexColor("#718096")
WHITE = colors.white

# Weekly bond-review workbook placed beside sample_data/ at the project root.
DEFAULT_BOND_REVIEW_PATH = os.path.join(
    PROJECT_ROOT, "sample_data", "Sinopac_Bond_Review_2025_1118.xlsx",
)

# --- CJK font registration -------------------------------------------------
# reportlab's built-in fonts (Helvetica, Times) have no Chinese glyphs, so
# any Traditional Chinese text (e.g. the LLM summaries) rendered with them
# comes out as solid black boxes ("tofu"). We embed a real CJK font instead.
_FONT_DIR = os.path.join(PROJECT_ROOT, "fonts")
_FONT_PATH = os.path.join(_FONT_DIR, "NotoSansTC-Regular.ttf")
_BOLD_FONT_PATH = os.path.join(_FONT_DIR, "NotoSansTC-Bold.ttf")

# Also support common local fonts when the project fonts/ directory was not
# copied. This is useful on the Windows machine where the report is run.
if not os.path.exists(_FONT_PATH):
    for regular_path, bold_path in (
        (r"C:\Windows\Fonts\msjh.ttc", r"C:\Windows\Fonts\msjhbd.ttc"),
        (r"C:\Windows\Fonts\msjh.ttf", r"C:\Windows\Fonts\msjhbd.ttf"),
        ("/System/Library/Fonts/PingFang.ttc",
         "/System/Library/Fonts/PingFang.ttc"),
    ):
        if os.path.exists(regular_path):
            _FONT_PATH = regular_path
            _BOLD_FONT_PATH = bold_path
            break
CJK_FONT_NAME = "Helvetica"  # fallback if the font file is missing
CJK_BOLD_NAME = "Helvetica-Bold"

if os.path.exists(_FONT_PATH):
    try:
        pdfmetrics.registerFont(TTFont("NotoSansTC", _FONT_PATH))
        CJK_FONT_NAME = "NotoSansTC"
        CJK_BOLD_NAME = "NotoSansTC"

        # Optional bold weight. Without this, <b> tags silently do nothing
        # (reportlab looks up the family's bold face; if the family isn't
        # registered it falls back to Helvetica-Bold => tofu on Chinese).
        if os.path.exists(_BOLD_FONT_PATH):
            try:
                pdfmetrics.registerFont(TTFont("NotoSansTC-Bold", _BOLD_FONT_PATH))
                CJK_BOLD_NAME = "NotoSansTC-Bold"
            except Exception:
                pass

        pdfmetrics.registerFontFamily(
            "NotoSansTC",
            normal="NotoSansTC",
            bold=CJK_BOLD_NAME,
            italic="NotoSansTC",
            boldItalic=CJK_BOLD_NAME,
        )
    except Exception:
        pass  # falls back to Helvetica; ASCII still renders, Chinese would tofu

def _cjk_styles():
    """Return a stylesheet where every style used in this document renders CJK."""
    styles = getSampleStyleSheet()
    for style_name in ("Title", "Heading1", "Heading2", "Heading3", "Normal", "Italic"):
        styles[style_name].fontName = CJK_FONT_NAME
    styles["Normal"].textColor = INK
    styles["Normal"].fontSize = 9.5
    styles["Normal"].leading = 15
    return styles


def _draw_page_frame(canvas, doc):
    """Draw a quiet running header, footer and page number."""
    canvas.saveState()
    page_width, page_height = doc.pagesize

    # Slim accent rule at the top, then a subtle report identifier.
    canvas.setFillColor(NAVY)
    canvas.rect(0, page_height - 7, page_width, 7, stroke=0, fill=1)
    canvas.setFillColor(MUTED)
    canvas.setFont(CJK_FONT_NAME, 7.5)
    canvas.drawString(doc.leftMargin, page_height - 28, config.REPORT_TITLE)

    # Footer rule and page number.
    canvas.setStrokeColor(BORDER)
    canvas.setLineWidth(0.5)
    canvas.line(doc.leftMargin, 35, page_width - doc.rightMargin, 35)
    canvas.setFillColor(MUTED)
    canvas.setFont(CJK_FONT_NAME, 7.5)
    canvas.drawString(doc.leftMargin, 22, "INTERNAL RISK REPORT")
    canvas.drawRightString(
        page_width - doc.rightMargin, 22, f"{canvas.getPageNumber():02d}"
    )
    canvas.restoreState()


def _section_header(title, style, index):
    """Return a compact numbered section banner."""
    number_style = ParagraphStyle(
        "SectionNumber", parent=style, fontName=CJK_BOLD_NAME,
        fontSize=10, leading=12, textColor=WHITE, alignment=1,
    )
    title_style = ParagraphStyle(
        "SectionTitle", parent=style, fontName=CJK_BOLD_NAME,
        fontSize=13, leading=18, textColor=NAVY,
    )
    banner = Table(
        [[Paragraph(f"{index:02d}", number_style),
          Paragraph(_inline(str(title)), title_style)]],
        colWidths=[0.48 * inch, None],
        hAlign="LEFT",
    )
    banner.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, 0), ACCENT),
        ("BACKGROUND", (1, 0), (1, 0), PALE_BLUE),
        ("BOX", (0, 0), (-1, -1), 0.6, BORDER),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (0, 0), 5),
        ("RIGHTPADDING", (0, 0), (0, 0), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 9),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 9),
        ("LEFTPADDING", (1, 0), (1, 0), 12),
        ("RIGHTPADDING", (1, 0), (1, 0), 10),
    ]))
    return banner


# --- Weekly bond-review workbook ------------------------------------------

def _is_number(value):
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def _find_bond_review_path(explicit_path=None):
    """Resolve the configured workbook, with a dated-file fallback."""
    if explicit_path:
        return explicit_path if os.path.exists(explicit_path) else None
    if os.path.exists(DEFAULT_BOND_REVIEW_PATH):
        return DEFAULT_BOND_REVIEW_PATH

    sample_dir = os.path.dirname(DEFAULT_BOND_REVIEW_PATH)
    candidates = sorted(
        glob.glob(os.path.join(sample_dir, "Sinopac_Bond_Review_*.xlsx")),
        reverse=True,
    )
    return candidates[0] if candidates else None


def _worksheet_case_insensitive(workbook, name):
    for sheet_name in workbook.sheetnames:
        if sheet_name.strip().lower() == name.strip().lower():
            return workbook[sheet_name]
    return None


def _extract_tflo_block(sheet, marker_text):
    """Read one TFLO block after its marker and return normalized rows."""
    marker_row = None
    for row in range(1, sheet.max_row + 1):
        value = sheet.cell(row, 1).value
        if value and marker_text.lower() in str(value).lower():
            marker_row = row
            break
    if marker_row is None:
        return []

    total_row = None
    for row in range(marker_row + 1, min(marker_row + 10, sheet.max_row) + 1):
        if str(sheet.cell(row, 1).value or "").strip().lower() == "total":
            total_row = row
            break
    if total_row is None:
        return []

    rows = []
    for row in range(total_row, sheet.max_row + 1):
        label = sheet.cell(row, 1).value
        if row > total_row and (label is None or str(label).strip() == ""):
            break
        volume = sheet.cell(row, 2).value
        if not label or not _is_number(volume):
            continue
        rows.append({
            "label": str(label).strip(),
            "volume": float(volume),
            "buys": float(sheet.cell(row, 3).value or 0),
            "sells": float(sheet.cell(row, 4).value or 0),
            "net": float(sheet.cell(row, 5).value or 0),
            "buy_sell_pct": float(sheet.cell(row, 6).value or 0),
        })
    return rows


def _extract_template_bonds(template_sheet):
    """Read Template's Bloomberg-backed Top 20 table when cached values exist."""
    if template_sheet is None:
        return []
    rows = []
    # Template!BT8:CF27 is the table shown on page 4 of 1118.pdf.
    for row in range(8, 28):
        ticker = template_sheet[f"BT{row}"].value
        volume = template_sheet[f"CC{row}"].value
        if (not ticker or str(ticker).startswith("#") or
                not _is_number(volume)):
            continue
        rows.append({
            "ticker": str(ticker).strip(),
            "issuer": str(template_sheet[f"BU{row}"].value or "").strip(),
            "coupon": template_sheet[f"BW{row}"].value,
            "maturity": template_sheet[f"BX{row}"].value,
            "yield": template_sheet[f"CA{row}"].value,
            "trades": template_sheet[f"CB{row}"].value,
            "volume": float(volume),
        })
    return sorted(rows, key=lambda item: item["volume"], reverse=True)


def _load_bond_review(workbook_path=None):
    """Return chart-ready data; failure never prevents the main PDF build."""
    path = _find_bond_review_path(workbook_path)
    if not path:
        return None
    try:
        workbook = load_workbook(path, data_only=True, read_only=False)
        tflo = _worksheet_case_insensitive(workbook, "TFLO")
        template = _worksheet_case_insensitive(workbook, "Template")
        if tflo is None:
            return None

        sectors = _extract_tflo_block(tflo, "TFLO->Sectors")
        maturities = _extract_tflo_block(tflo, "TFLO->Maturity")
        ratings = _extract_tflo_block(tflo, "TFLO->Rating")
        tickers = _extract_tflo_block(tflo, "TFLO->Tickers")
        bonds = _extract_template_bonds(template)

        sector_rows = [r for r in sectors if r["label"].lower() != "total"]
        maturity_rows = [
            r for r in maturities
            if r["label"].lower() not in ("total", "matured")
        ]
        rating_labels = {"investment grade", "high yield", "not rated"}
        rating_rows = [
            r for r in ratings if r["label"].lower() in rating_labels
        ]
        ticker_rows = [r for r in tickers if r["label"].lower() != "total"]
        total = next(
            (r for r in sectors if r["label"].lower() == "total"), None
        )

        as_of = ""
        for row in range(1, min(tflo.max_row, 8) + 1):
            for col in range(1, min(tflo.max_column, 14) + 1):
                value = tflo.cell(row, col).value
                if value and "as of" in str(value).lower():
                    as_of = str(value).replace("As of", "").strip()
                    break
            if as_of:
                break

        if not any((sector_rows, maturity_rows, rating_rows, ticker_rows, bonds)):
            return None
        return {
            "path": path,
            "filename": os.path.basename(path),
            "as_of": as_of,
            "total": total,
            "sectors": sorted(
                sector_rows, key=lambda item: item["volume"], reverse=True
            ),
            "maturities": maturity_rows,
            "ratings": rating_rows,
            "tickers": sorted(
                ticker_rows, key=lambda item: item["volume"], reverse=True
            ),
            "bonds": bonds,
        }
    except Exception:
        return None


def _short_label(value, max_chars=20):
    value = str(value)
    return value if len(value) <= max_chars else value[:max_chars - 3] + "..."


def _ranked_bar_drawing(title, rows, width=244, height=174, limit=8,
                        value_divisor=1_000_000, unit="USD bn"):
    """Create a compact, vector-based horizontal ranking chart."""
    drawing = Drawing(width, height)
    drawing.add(String(
        0, height - 14, title, fontName=CJK_BOLD_NAME,
        fontSize=9.5, fillColor=NAVY,
    ))
    drawing.add(Line(0, height - 21, width, height - 21,
                     strokeColor=BORDER, strokeWidth=0.6))

    rows = list(rows or [])[:limit]
    if not rows:
        drawing.add(String(
            0, height / 2, "無可用資料", fontName=CJK_FONT_NAME,
            fontSize=8, fillColor=MUTED,
        ))
        return drawing

    values = [max(float(row["volume"]) / value_divisor, 0) for row in rows]
    maximum = max(values) or 1
    top = height - 39
    row_height = max((height - 48) / len(rows), 11)
    label_width = 73
    value_width = 34
    bar_width = width - label_width - value_width - 7

    for index, (row, value) in enumerate(zip(rows, values)):
        y = top - index * row_height
        label = row.get("ticker") or row.get("label") or ""
        drawing.add(String(
            0, y, _short_label(label, 17), fontName=CJK_FONT_NAME,
            fontSize=6.8, fillColor=INK,
        ))
        drawing.add(Rect(
            label_width, y - 1, bar_width, 6,
            fillColor=PALE_BLUE, strokeColor=None,
        ))
        drawing.add(Rect(
            label_width, y - 1, bar_width * value / maximum, 6,
            fillColor=ACCENT if index else NAVY, strokeColor=None,
        ))
        drawing.add(String(
            width, y, f"{value:,.1f}", fontName=CJK_FONT_NAME,
            fontSize=6.8, fillColor=SLATE, textAnchor="end",
        ))
    drawing.add(String(
        width, 1, unit, fontName=CJK_FONT_NAME,
        fontSize=6.2, fillColor=MUTED, textAnchor="end",
    ))
    return drawing


def _metric_card(label, value, style):
    label_style = ParagraphStyle(
        "MetricLabel", parent=style, fontName=CJK_FONT_NAME,
        fontSize=6.8, leading=9, textColor=MUTED,
    )
    value_style = ParagraphStyle(
        "MetricValue", parent=style, fontName=CJK_BOLD_NAME,
        fontSize=13, leading=16, textColor=NAVY,
    )
    return [Paragraph(_inline(label), label_style),
            Paragraph(_inline(value), value_style)]


def _bond_review_flowables(snapshot, styles, section_index, doc_width):
    """Build the data-driven section inspired by 1118.pdf."""
    story = [PageBreak(), _section_header(
        "債券市場數據總覽", styles["Heading1"], section_index
    ), Spacer(1, 10)]

    total = snapshot.get("total") or {}
    sectors = snapshot.get("sectors") or []
    tickers = snapshot.get("tickers") or []
    bonds = snapshot.get("bonds") or []
    total_volume = float(total.get("volume", 0)) / 1_000_000
    net_volume = float(total.get("net", 0)) / 1_000_000
    largest_sector = sectors[0] if sectors else None
    most_active = bonds[0] if bonds else (tickers[0] if tickers else None)
    active_name = "-"
    if most_active:
        active_name = most_active.get("ticker") or most_active.get("label") or "-"

    cards = Table([[
        _metric_card("公司債週交易量", f"{total_volume:,.1f} USD bn", styles["Normal"]),
        _metric_card("Dealer-to-client 淨額", f"{net_volume:+,.1f} USD bn", styles["Normal"]),
        _metric_card("最大交易產業", largest_sector["label"] if largest_sector else "-", styles["Normal"]),
        _metric_card("最活躍債券／Ticker", active_name, styles["Normal"]),
    ]], colWidths=[doc_width / 4.0] * 4)
    cards.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), PALE_GREY),
        ("BOX", (0, 0), (-1, -1), 0.6, BORDER),
        ("INNERGRID", (0, 0), (-1, -1), 0.4, BORDER),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 9),
        ("RIGHTPADDING", (0, 0), (-1, -1), 7),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))
    story.extend([cards, Spacer(1, 7)])

    meta_style = ParagraphStyle(
        "BondDataMeta", parent=styles["Normal"], fontName=CJK_FONT_NAME,
        fontSize=7, leading=10, textColor=MUTED,
    )
    as_of = f"｜資料時間：{snapshot['as_of']}" if snapshot.get("as_of") else ""
    story.extend([
        Paragraph(
            f"資料來源：{_inline(snapshot['filename'])}{_inline(as_of)}",
            meta_style,
        ),
        Spacer(1, 8),
    ])

    ranking_rows = bonds if bonds else tickers
    ranking_title = ("投資等級債券交易量排行" if bonds
                     else "活躍發行人交易量排行")
    ranking_divisor = 1_000 if bonds else 1_000_000
    ranking_unit = "USD mn" if bonds else "USD bn"
    charts = [
        _ranked_bar_drawing(ranking_title, ranking_rows, limit=10,
                            value_divisor=ranking_divisor, unit=ranking_unit),
        _ranked_bar_drawing("產業交易量排行", sectors, limit=8),
        _ranked_bar_drawing("天期交易量分布", snapshot.get("maturities"), limit=7),
        _ranked_bar_drawing("信評交易量分布", snapshot.get("ratings"), limit=4),
    ]
    chart_grid = Table(
        [[charts[0], charts[1]], [charts[2], charts[3]]],
        colWidths=[doc_width / 2.0] * 2,
        rowHeights=[180, 180],
    )
    chart_grid.setStyle(TableStyle([
        ("BOX", (0, 0), (-1, -1), 0.5, BORDER),
        ("INNERGRID", (0, 0), (-1, -1), 0.5, BORDER),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    story.append(chart_grid)

    # If Bloomberg formulas have cached values, include the bond-level Top 10
    # table from Template. Otherwise the Ticker ranking above is the fallback.
    if bonds:
        story.extend([Spacer(1, 12), Paragraph(
            "投資等級債券交易量 Top 10", styles["Heading3"]
        )])
        header = ["Ticker", "發行人", "票息", "天期", "殖利率", "筆數", "總量(M)"]
        table_rows = [header]
        for bond in bonds[:10]:
            table_rows.append([
                bond["ticker"], _short_label(bond["issuer"], 16),
                bond["coupon"], bond["maturity"], bond["yield"],
                bond["trades"], bond["volume"],
            ])
        cell_style = ParagraphStyle(
            "BondCell", parent=styles["Normal"], fontName=CJK_FONT_NAME,
            fontSize=6.5, leading=8,
        )
        data = [[Paragraph(_inline(str(v or "")), cell_style) for v in row]
                for row in table_rows]
        bond_table = Table(data, repeatRows=1, colWidths=[
            0.82 * inch, 1.25 * inch, 0.48 * inch, 0.45 * inch,
            0.52 * inch, 0.42 * inch, 0.72 * inch,
        ])
        bond_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), NAVY),
            ("TEXTCOLOR", (0, 0), (-1, 0), WHITE),
            ("BOX", (0, 0), (-1, -1), 0.5, BORDER),
            ("LINEBELOW", (0, 1), (-1, -1), 0.3, BORDER),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(bond_table)
    return story


# --- Markdown -> flowables --------------------------------------------------

_BULLET_RE = re.compile(r"^\s*[-*•·‧–]\s+(.*)$")
_ORDERED_RE = re.compile(r"^\s*(\d+)\s*[.)、]\s+(.*)$")
_HEADING_RE = re.compile(r"^\s*(#{1,6})\s*(.+?)\s*#*$")
_HR_RE = re.compile(r"^\s*([-*_])\s*(?:\1\s*){2,}$")
_TABLE_SEP_RE = re.compile(r"^\s*\|?[\s:|-]+\|[\s:|-]*$")
_LEAD_WS_RE = re.compile(r"^(\s*)")

_BOLD_RE = re.compile(r"\*\*(.+?)\*\*|__(.+?)__")
_ITALIC_RE = re.compile(r"(?<![\*\w])\*(?!\s)([^*]+?)(?<!\s)\*(?!\*)")
_CODE_RE = re.compile(r"`([^`]+)`")
_LINK_RE = re.compile(r"\[([^\]]+)\]\(([^)]+)\)")


def _inline(text: str) -> str:
    """
    Inline Markdown -> reportlab mini-HTML.

    The escaping matters: reportlab 4.x tolerates a bare '&' or '<GO>', but a
    stray unclosed tag like '<b>' or '<br' raises and kills the whole build.
    Escaping first means source text can never be mistaken for markup — only
    the tags we deliberately generate below survive.
    """
    text = (text.replace("&", "&amp;")
                .replace("<", "&lt;")
                .replace(">", "&gt;"))
    text = _BOLD_RE.sub(lambda m: "<b>%s</b>" % (m.group(1) or m.group(2)), text)
    text = _ITALIC_RE.sub(r"<i>\1</i>", text)
    text = _CODE_RE.sub(r"<font face='Courier'>\1</font>", text)
    text = _LINK_RE.sub(r"\1", text)  # keep link text, drop the URL parens
    return text.strip()


def _is_table_row(line: str) -> bool:
    s = line.strip()
    return s.startswith("|") and s.count("|") >= 2


def _split_row(line: str) -> list:
    return [c.strip() for c in line.strip().strip("|").split("|")]


def _render_markdown(text, style):
    """
    Turn Markdown text (or an already-clean list of bullet strings) into
    reportlab flowables. Public behaviour matches the old
    _bullets_to_flowables(text, style) signature.
    """
    bullet_style = ParagraphStyle(
        "MDBullet", parent=style,
        leftIndent=(getattr(style, "leftIndent", 12) or 12),
        bulletIndent=max((getattr(style, "leftIndent", 12) or 12) - 10, 0),
        spaceAfter=0,
    )
    sub_bullet_style = ParagraphStyle(
        "MDSubBullet", parent=bullet_style,
        leftIndent=bullet_style.leftIndent + 14,
        bulletIndent=bullet_style.bulletIndent + 14,
    )
    body_style = ParagraphStyle(
        "MDBody", parent=style, leftIndent=0, bulletIndent=0, spaceAfter=0,
    )
    heading_style = ParagraphStyle(
        "MDHeading", parent=style, fontName=CJK_BOLD_NAME,
        fontSize=(style.fontSize or 10) + 1.5,
        leading=(style.leading or 14) + 3,
        textColor=NAVY,
        spaceBefore=9, spaceAfter=4, leftIndent=0, bulletIndent=0,
    )

    # Upstream may already hand us a clean list of bullets (llm.as_bullets).
    if isinstance(text, (list, tuple)):
        out = []
        for item in text:
            item = str(item).strip()
            if item:
                out.append(Paragraph(_inline(item), bullet_style, bulletText="•"))
                out.append(Spacer(1, 3))
        return out

    text = (text or "").strip()
    if not text:
        return []

    flowables = []
    table_buf = []

    def flush_table():
        if not table_buf:
            return
        rows = [_split_row(r) for r in table_buf if not _TABLE_SEP_RE.match(r)]
        table_buf.clear()
        if not rows:
            return
        width = max(len(r) for r in rows)
        cell_style = ParagraphStyle(
            "MDCell", parent=body_style,
            fontSize=max((style.fontSize or 10) - 1.5, 7),
            leading=max((style.leading or 14) - 3, 8),
        )
        header_cell_style = ParagraphStyle(
            "MDHeaderCell", parent=cell_style, fontName=CJK_BOLD_NAME,
            textColor=WHITE,
        )
        data = []
        for row_index, row in enumerate(rows):
            padded_row = row + [""] * (width - len(row))
            row_style = header_cell_style if row_index == 0 else cell_style
            data.append([Paragraph(_inline(c), row_style) for c in padded_row])
        tbl = Table(data, hAlign="LEFT")
        table_commands = [
            ("BOX", (0, 0), (-1, -1), 0.6, BORDER),
            ("LINEBELOW", (0, 0), (-1, 0), 0.8, ACCENT),
            ("LINEBELOW", (0, 1), (-1, -1), 0.35, BORDER),
            ("BACKGROUND", (0, 0), (-1, 0), NAVY),
            ("TEXTCOLOR", (0, 0), (-1, 0), WHITE),
            ("FONTNAME", (0, 0), (-1, 0), CJK_BOLD_NAME),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 7),
            ("RIGHTPADDING", (0, 0), (-1, -1), 7),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]
        # Very light zebra striping improves row tracking without visual noise.
        for row_index in range(2, len(data), 2):
            table_commands.append(
                ("BACKGROUND", (0, row_index), (-1, row_index), PALE_GREY)
            )
        tbl.setStyle(TableStyle(table_commands))
        flowables.append(tbl)
        flowables.append(Spacer(1, 6))

    for raw_line in text.splitlines():
        line = raw_line.rstrip()

        if _is_table_row(line):
            table_buf.append(line)
            continue
        flush_table()

        if not line.strip():
            continue
        if _HR_RE.match(line):
            flowables.append(Spacer(1, 4))
            flowables.append(HRFlowable(
                width="100%", thickness=0.5, color=BORDER,
                spaceBefore=2, spaceAfter=5,
            ))
            continue

        m = _HEADING_RE.match(line)
        if m:
            flowables.append(Paragraph(_inline(m.group(2)), heading_style))
            continue

        indent = len(_LEAD_WS_RE.match(raw_line).group(1))

        m = _BULLET_RE.match(line)
        if m:
            nested = indent >= 2
            flowables.append(Paragraph(
                _inline(m.group(1)),
                sub_bullet_style if nested else bullet_style,
                bulletText="–" if nested else "•",
            ))
            flowables.append(Spacer(1, 3))
            continue

        m = _ORDERED_RE.match(line)
        if m:
            nested = indent >= 2
            flowables.append(Paragraph(
                _inline(m.group(2)),
                sub_bullet_style if nested else bullet_style,
                bulletText="%s." % m.group(1),
            ))
            flowables.append(Spacer(1, 3))
            continue

        flowables.append(Paragraph(_inline(line), body_style))
        flowables.append(Spacer(1, 4))

    flush_table()
    return flowables


def _bullets_to_flowables(text, style):
    """Backwards-compatible name; now Markdown-aware."""
    return _render_markdown(text, style)


def build_pdf(
    target_date: str,
    sections: list,
    output_path: str,
    chart_paths: list = None,
    extracted_reports: list = None,
    bond_review_path: str = None,
):
    """
    sections: list of dicts, each:
        {
            "title": str,           # e.g. "第一區：重點債市新聞與數據評析"
            "body": str | list,     # LLM-synthesized markdown, or list of bullets
            "sources": list[str],   # source filenames/subjects, for traceability
        }
    bond_review_path: optional explicit xlsx path. When omitted, this module
    looks for sample_data/Sinopac_Bond_Review_2025_1118.xlsx at the project
    root (see DEFAULT_BOND_REVIEW_PATH above).
    """
    styles = _cjk_styles()
    bullet_style = ParagraphStyle(
        "Bullet", parent=styles["Normal"], leftIndent=14, bulletIndent=2,
        fontSize=9.5, leading=15, fontName=CJK_FONT_NAME,
        textColor=INK, spaceAfter=1,
    )
    source_style = ParagraphStyle(
        "Sources", parent=styles["Normal"], fontName=CJK_FONT_NAME,
        fontSize=7.5, textColor=MUTED, leading=11,
        backColor=PALE_GREY, borderColor=BORDER, borderWidth=0.5,
        borderPadding=(6, 8, 6, 8),
    )

    title_style = ParagraphStyle(
        "ReportTitle", parent=styles["Title"], fontName=CJK_BOLD_NAME,
        fontSize=23, leading=30, textColor=WHITE, alignment=0,
        spaceAfter=0,
    )
    date_style = ParagraphStyle(
        "ReportDate", parent=styles["Normal"], fontName=CJK_FONT_NAME,
        fontSize=10, leading=14, textColor=colors.HexColor("#DCE8F1"),
    )
    label_style = ParagraphStyle(
        "ReportLabel", parent=styles["Normal"], fontName=CJK_BOLD_NAME,
        fontSize=7.5, leading=10, textColor=colors.HexColor("#BFD3E2"),
        tracking=0.8,
    )

    doc = SimpleDocTemplate(
        output_path,
        pagesize=letter,
        leftMargin=0.68 * inch,
        rightMargin=0.68 * inch,
        topMargin=0.62 * inch,
        bottomMargin=0.62 * inch,
        title=config.REPORT_TITLE,
        author="Risk Management",
    )
    story = []

    # --- Title ---
    title_panel = Table(
        [[Paragraph("DAILY MARKET &amp; RISK DIGEST", label_style)],
         [Paragraph(_inline(config.REPORT_TITLE), title_style)],
         [Paragraph(_inline(target_date), date_style)]],
        colWidths=[doc.width],
        hAlign="LEFT",
    )
    title_panel.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), NAVY),
        ("LEFTPADDING", (0, 0), (-1, -1), 22),
        ("RIGHTPADDING", (0, 0), (-1, -1), 22),
        ("TOPPADDING", (0, 0), (0, 0), 18),
        ("BOTTOMPADDING", (0, 0), (0, 0), 4),
        ("TOPPADDING", (0, 1), (0, 1), 2),
        ("BOTTOMPADDING", (0, 1), (0, 1), 8),
        ("TOPPADDING", (0, 2), (0, 2), 2),
        ("BOTTOMPADDING", (0, 2), (0, 2), 18),
        ("LINEBELOW", (0, 1), (0, 1), 1.2, ACCENT),
    ]))
    story.append(title_panel)
    story.append(Spacer(1, 20))

    # --- Sections ---
    for section_index, section in enumerate(sections, start=1):
        story.append(KeepTogether([
            _section_header(section["title"], styles["Heading1"], section_index),
            Spacer(1, 10),
        ]))
        story.extend(_bullets_to_flowables(section["body"], bullet_style))
        sources = section.get("sources") or []
        if sources:
            story.append(Spacer(1, 7))
            story.append(Paragraph(
                "<b>來源</b>  " + " · ".join(_inline(str(s)) for s in sources),
                source_style))
        story.append(Spacer(1, 22))

    next_section_index = len(sections) + 1

    # --- Data-driven weekly bond charts (from Template / TFLO) ---
    bond_snapshot = _load_bond_review(bond_review_path)
    if bond_snapshot:
        story.extend(_bond_review_flowables(
            bond_snapshot, styles, next_section_index, doc.width
        ))
        next_section_index += 1

    # --- Charts (currently unused; see config.ENABLE_CHARTS) ---
    if chart_paths:
        story.append(PageBreak())
        story.append(_section_header(
            "Key Metrics", styles["Heading1"], next_section_index
        ))
        next_section_index += 1
        story.append(Spacer(1, 12))
        for path in chart_paths:
            # kind="proportional" so charts aren't stretched to a fixed box.
            story.append(Image(path, width=6.2 * inch, height=3.5 * inch,
                               kind="proportional"))
            story.append(Spacer(1, 10))

    # --- Charts embedded in the source reports themselves (PDF mode only) ---
    reports_with_images = [
        r for r in (extracted_reports or []) if getattr(r, "images", None)
    ]
    if reports_with_images:
        story.append(PageBreak())
        story.append(_section_header(
            "Charts from Source Reports", styles["Heading1"],
            next_section_index,
        ))
        story.append(Spacer(1, 10))
        story.append(Paragraph(
            "These are charts embedded directly in the original PDFs "
            "(not reconstructed) — the underlying data wasn't available "
            "as an extractable table.", styles["Italic"]
        ))
        story.append(Spacer(1, 8))
        for r in reports_with_images:
            story.append(Paragraph(r.filename, styles["Heading3"]))
            for img in r.images[:MAX_SOURCE_IMAGES_PER_FILE]:
                try:
                    img_flowable = Image(io.BytesIO(img["bytes"]),
                                         width=5.5 * inch, height=3.2 * inch,
                                         kind="proportional")
                    story.append(img_flowable)
                    story.append(Spacer(1, 8))
                except Exception:
                    continue

    doc.build(story, onFirstPage=_draw_page_frame, onLaterPages=_draw_page_frame)
    return output_path
